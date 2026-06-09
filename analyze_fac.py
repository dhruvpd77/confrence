# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('fac data.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print('Size:', ws.max_row, 'x', ws.max_column)

# Find header row and columns
for row in range(1, 6):
    vals = []
    for col in range(1, min(ws.max_column + 1, 30)):
        v = ws.cell(row, col).value
        if v:
            vals.append(f'C{col}={repr(str(v)[:60])}')
    print(f'Row {row}: {vals}')

# Search for mobile column
print('\nSearching mobile/name columns...')
for col in range(1, ws.max_column + 1):
    for row in range(1, 5):
        v = str(ws.cell(row, col).value or '').lower()
        if 'mobile' in v or 'phone' in v or 'faculty member' in v:
            print(f'  Row {row} Col {col}: {ws.cell(row, col).value}')

# Sample data rows
print('\nSample faculty rows:')
name_col = None
mobile_col = None
for col in range(1, ws.max_column + 1):
    h = str(ws.cell(1, col).value or '') + str(ws.cell(2, col).value or '')
    if 'faculty member' in h.lower() or 'name of the faculty' in h.lower():
        name_col = col
    if 'mobile' in h.lower():
        mobile_col = col

for col in range(1, ws.max_column + 1):
    for row in range(1, 4):
        v = str(ws.cell(row, col).value or '')
        if 'mobile' in v.lower():
            mobile_col = col
        if 'faculty member' in v.lower():
            name_col = col

print(f'name_col={name_col}, mobile_col={mobile_col}')

if name_col and mobile_col:
    for row in range(3, 12):
        print(f'  {ws.cell(row, name_col).value} | {ws.cell(row, mobile_col).value}')
