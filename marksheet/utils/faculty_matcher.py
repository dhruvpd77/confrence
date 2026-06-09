import re
from difflib import SequenceMatcher
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook


TITLES = re.compile(
    r'^(DR\.?|MR\.?|MS\.?|MRS\.?|PROF\.?|PROFESSOR)\s+',
    re.I,
)


def normalize_name(name):
    text = str(name or '').strip().upper()
    while True:
        cleaned = TITLES.sub('', text).strip()
        if cleaned == text:
            break
        text = cleaned
    text = re.sub(r'[.\-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def name_tokens(name):
    return [t for t in normalize_name(name).split() if t]


def _similar_token(a, b, threshold=0.82):
    if a == b:
        return True
    return SequenceMatcher(None, a, b).ratio() >= threshold


def format_mobile(value):
    if value is None:
        return None
    if isinstance(value, float):
        if value == int(value):
            value = int(value)
    text = re.sub(r'\D', '', str(value))
    return text if len(text) >= 10 else None


class FacultyMatcher:
    @staticmethod
    def _default_path():
        return Path(settings.FACULTY_DATA_FILE)

    def _load(self):
        workbook = load_workbook(self.file_path, data_only=True)
        worksheet = workbook.active

        name_col = 2
        mobile_col = 22
        email_col = 23

        for row in range(3, worksheet.max_row + 1):
            name = worksheet.cell(row, name_col).value
            if not name:
                continue

            mobile = format_mobile(worksheet.cell(row, mobile_col).value)
            email_raw = worksheet.cell(row, email_col).value
            email = str(email_raw).strip().lower() if email_raw else ''
            if email and '@' not in email:
                email = ''

            normalized = normalize_name(name)
            tokens = name_tokens(name)
            if not normalized or not tokens:
                continue

            record = {
                'name': str(name).strip(),
                'normalized': normalized,
                'tokens': tokens,
                'mobile': mobile or '',
                'email': email,
                'first': tokens[0],
                'last': tokens[-1],
            }
            self.records.append(record)
            if mobile:
                self.by_normalized[normalized] = mobile
            if email:
                self.by_email_normalized[normalized] = email

        workbook.close()

    def __init__(self, file_path=None):
        self.file_path = Path(file_path) if file_path else self._default_path()
        self.records = []
        self.by_normalized = {}
        self.by_email_normalized = {}
        if self.file_path and self.file_path.exists():
            self._load()

    def _find_record_field(self, name, field):
        if not name or not self.records:
            return None

        normalized = normalize_name(name)
        lookup = self.by_email_normalized if field == 'email' else self.by_normalized
        if normalized in lookup:
            return lookup[normalized]

        tokens = name_tokens(name)
        if not tokens:
            return None

        first, last = tokens[0], tokens[-1]
        candidates = [r for r in self.records if r['first'] == first and r['last'] == last]
        if len(candidates) == 1:
            return candidates[0][field] or None
        if len(candidates) > 1:
            token_matches = [r for r in candidates if all(t in r['tokens'] for t in tokens)]
            if len(token_matches) == 1:
                return token_matches[0][field] or None

        subset_matches = [r for r in self.records if all(t in r['tokens'] for t in tokens)]
        if len(subset_matches) == 1:
            return subset_matches[0][field] or None

        last_matches = [r for r in self.records if r['last'] == last]
        if len(last_matches) == 1 and first == last_matches[0]['first']:
            return last_matches[0][field] or None

        first_last_in_faculty = [
            r for r in self.records
            if r['first'] == first and last in r['tokens']
        ]
        if len(first_last_in_faculty) == 1:
            return first_last_in_faculty[0][field] or None

        if len(tokens) >= 2:
            reversed_matches = [
                r for r in self.records
                if r['last'] == tokens[0] and r['first'] == tokens[1]
            ]
            if len(reversed_matches) == 1:
                return reversed_matches[0][field] or None

        fuzzy_matches = [
            r for r in self.records
            if r['first'] == first and _similar_token(r['last'], last)
        ]
        if len(fuzzy_matches) == 1:
            return fuzzy_matches[0][field] or None

        return None

    def find_email(self, name):
        return self._find_record_field(name, 'email')

    def find_contact(self, name):
        return {
            'phone': self.find_mobile(name) or '',
            'email': self.find_email(name) or '',
        }

    def find_mobile(self, name):
        value = self._find_record_field(name, 'mobile')
        return value or None

    def format_with_mobile(self, name):
        if not name:
            return name
        mobile = self.find_mobile(name)
        if mobile:
            return f'{name}\n({mobile})'
        return name
