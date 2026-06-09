import re

from openpyxl import load_workbook

from marksheet.models import TrackDuty
from marksheet.utils.excel_parser import _extract_day_info


def _parse_room_track_cell(value):
    text = str(value or '').strip()
    if ' - ' in text:
        room, track_session = text.split(' - ', 1)
        return room.strip(), track_session.strip()
    if '|' in text:
        parts = [p.strip() for p in text.split('|')]
        return parts[0], parts[1] if len(parts) > 1 else ''
    return text, ''


def _find_header_columns(worksheet, header_row=3):
    mapping = {}
    for col in range(1, worksheet.max_column + 1):
        val = str(worksheet.cell(header_row, col).value or '').strip().upper()
        if 'ROOM' in val:
            mapping['room'] = col
        elif val == 'VERIFIER' or val.startswith('VERIFIER'):
            mapping['verifier'] = col
        elif val == 'TRACK' or val.startswith('TRACK'):
            mapping['track'] = col
    return mapping


def parse_verifier_assignments(file_path):
    """Parse track duty Excel with VERIFIER column (Day 1 / Day 2 sheets)."""
    workbook = load_workbook(file_path, data_only=True)
    assignments = []

    for sheet_name in workbook.sheetnames:
        day_num, day_label = _extract_day_info(sheet_name)
        if day_num is None:
            if sheet_name.lower().startswith('day'):
                try:
                    day_num = int(re.search(r'\d+', sheet_name).group())
                except (AttributeError, ValueError):
                    continue
            else:
                continue

        worksheet = workbook[sheet_name]
        cols = _find_header_columns(worksheet)
        if 'verifier' not in cols:
            cols = {'room': 1, 'verifier': 2, 'track': 3}

        for row in range(4, worksheet.max_row + 1):
            room_col = cols.get('room', 1)
            verifier_col = cols.get('verifier', 2)

            room_raw = worksheet.cell(row, room_col).value
            verifier = worksheet.cell(row, verifier_col).value
            verifier = str(verifier or '').strip()
            if not verifier:
                continue

            room, track_session = _parse_room_track_cell(room_raw)
            if not track_session and 'track' in cols:
                track_name = worksheet.cell(row, cols['track']).value
                track_session = str(track_name or '').strip()

            if not room and not track_session:
                continue

            assignments.append({
                'day': day_num,
                'day_label': day_label,
                'room': room,
                'track_session': track_session,
                'verifier': verifier,
            })

    workbook.close()
    return assignments


def apply_verifier_assignments(schedule, assignments):
    updated = 0
    for item in assignments:
        duties = TrackDuty.objects.filter(
            schedule=schedule,
            day=item['day'],
        )
        if item['track_session']:
            duties = duties.filter(track_session=item['track_session'])
        if item['room']:
            duties = duties.filter(room=item['room'])

        count = duties.update(verifier=item['verifier'])
        updated += count

        if count == 0 and item['track_session']:
            duties = TrackDuty.objects.filter(
                schedule=schedule,
                day=item['day'],
                track_session__icontains=item['track_session'].split('_')[0],
            )
            if item['room']:
                duties = duties.filter(room=item['room'])
            updated += duties.update(verifier=item['verifier'])

    return updated
