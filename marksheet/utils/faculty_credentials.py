import random
import re

from django.contrib.auth import get_user_model

from marksheet.models import FacultyProfile, TrackDuty
from marksheet.utils.faculty_matcher import FacultyMatcher, name_tokens, normalize_name

User = get_user_model()

# Faculty logins and marks entry are only for Moderator-2 (Entry)
MODERATOR_2_FIELD = 'moderator_2'
MODERATOR_2_LABEL = 'Moderator-2 (Entry)'
LOGIN_STAFF_FIELDS = [MODERATOR_2_FIELD]


def _generate_password():
    return f'{random.randint(1000, 9999):04d}'


def _generate_username(display_name, used_usernames):
    tokens = name_tokens(display_name)
    if len(tokens) >= 2:
        base = f'{tokens[0].lower()}.{tokens[-1].lower()}'
    elif tokens:
        base = tokens[0].lower()
    else:
        base = 'faculty'
    base = re.sub(r'[^a-z0-9.]', '', base)[:24] or 'faculty'
    username = base
    counter = 1
    while username in used_usernames or User.objects.filter(username=username).exists():
        suffix = str(counter)
        username = f'{base[:24 - len(suffix)]}{suffix}'
        counter += 1
    used_usernames.add(username)
    return username


def _duty_verifier_lookup(schedule):
    """Previous verifier assignments keyed by (day, track_session) and (day, track_session, room)."""
    by_session = {}
    by_room = {}
    for duty in TrackDuty.objects.filter(schedule=schedule):
        if not duty.verifier:
            continue
        by_session[(duty.day, duty.track_session)] = duty.verifier
        by_room[(duty.day, duty.track_session, duty.room)] = duty.verifier
    return by_session, by_room


def sync_track_duties(schedule, duty_by_day, preserve_verifier=True):
    by_session, by_room = ({}, {})
    if preserve_verifier:
        by_session, by_room = _duty_verifier_lookup(schedule)

    TrackDuty.objects.filter(schedule=schedule).delete()
    duties = []
    for day_num, day_info in duty_by_day.items():
        for track in day_info['tracks']:
            verifier = ''
            if preserve_verifier:
                key_room = (day_num, track['track_session'], track['room'])
                key_session = (day_num, track['track_session'])
                verifier = by_room.get(key_room) or by_session.get(key_session, '')

            duties.append(TrackDuty(
                schedule=schedule,
                day=day_num,
                day_label=day_info.get('day_label', ''),
                room=track['room'],
                track_session=track['track_session'],
                track_name=track['track_name'],
                session_chair=track['session_chair'],
                track_coordinator=track['track_coordinator'],
                moderator_1=track['moderator_1'],
                moderator_2=track['moderator_2'],
                moderator_3=track['moderator_3'],
                verifier=verifier,
            ))
    TrackDuty.objects.bulk_create(duties)
    return duties


def sync_faculty_users(schedule, duty_by_day, preserve_existing=True):
    faculty_matcher = FacultyMatcher()
    unique_names = {}

    for day_info in duty_by_day.values():
        for track in day_info['tracks']:
            for field in LOGIN_STAFF_FIELDS:
                name = str(track.get(field) or '').strip()
                if not name:
                    continue
                norm = normalize_name(name)
                if norm and norm not in unique_names:
                    unique_names[norm] = name

    old_profiles = list(
        FacultyProfile.objects.filter(schedule=schedule).select_related('user')
    )
    old_by_norm = {p.normalized_name: p for p in old_profiles}
    needed_norms = set(unique_names.keys())

    if preserve_existing and old_profiles:
        kept = []
        for norm in needed_norms:
            if norm in old_by_norm:
                profile = old_by_norm[norm]
                display_name = unique_names[norm]
                if profile.display_name != display_name:
                    profile.display_name = display_name
                    profile.user.first_name = display_name[:30]
                    profile.user.save(update_fields=['first_name'])
                    profile.save(update_fields=['display_name'])
                kept.append(profile)

        for norm, profile in old_by_norm.items():
            if norm not in needed_norms:
                user_id = profile.user_id
                profile.delete()
                User.objects.filter(id=user_id, is_superuser=False).delete()

        used_norms = {p.normalized_name for p in kept}
        used_usernames = set(User.objects.values_list('username', flat=True))
        for norm, display_name in sorted(unique_names.items(), key=lambda x: x[1]):
            if norm in used_norms:
                continue
            username = _generate_username(display_name, used_usernames)
            password = _generate_password()
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=display_name[:30],
            )
            phone = faculty_matcher.find_mobile(display_name) or ''
            profile = FacultyProfile.objects.create(
                user=user,
                display_name=display_name,
                normalized_name=norm,
                plain_password=password,
                phone=phone or '',
                schedule=schedule,
            )
            kept.append(profile)
        return kept

    old_user_ids = [p.user_id for p in old_profiles]
    FacultyProfile.objects.filter(schedule=schedule).delete()
    User.objects.filter(id__in=old_user_ids, is_superuser=False).delete()

    used_usernames = set(User.objects.values_list('username', flat=True))
    created = []

    for norm, display_name in sorted(unique_names.items(), key=lambda x: x[1]):
        username = _generate_username(display_name, used_usernames)
        password = _generate_password()
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=display_name[:30],
        )
        phone = faculty_matcher.find_mobile(display_name) or ''
        profile = FacultyProfile.objects.create(
            user=user,
            display_name=display_name,
            normalized_name=norm,
            plain_password=password,
            phone=phone or '',
            schedule=schedule,
        )
        created.append(profile)

    return created


def names_match(name_a, name_b):
    if not name_a or not name_b:
        return False
    return normalize_name(name_a) == normalize_name(name_b)


def faculty_matches_name(profile, name):
    return names_match(profile.display_name, name) or profile.normalized_name == normalize_name(name)


def get_faculty_duties(profile, schedule=None, day=None):
    if schedule is None:
        schedule = profile.schedule
    if not schedule:
        return TrackDuty.objects.none()

    duties = TrackDuty.objects.filter(schedule=schedule)
    if day:
        duties = duties.filter(day=day)

    matched = []
    for duty in duties:
        if faculty_matches_name(profile, duty.moderator_2):
            matched.append({'duty': duty, 'roles': [MODERATOR_2_LABEL]})
    return matched


def get_moderator2_profiles(schedule):
    if not schedule:
        return FacultyProfile.objects.none()
    return FacultyProfile.objects.filter(schedule=schedule).select_related('user').order_by('display_name')


def get_faculty_track_keys(profile, schedule=None, day=None):
    keys = set()
    for item in get_faculty_duties(profile, schedule=schedule, day=day):
        duty = item['duty']
        keys.add((duty.day, duty.track_session))
    return keys


def generate_credentials_workbook(profiles):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Faculty Credentials'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    matcher = FacultyMatcher()

    headers = [
        'Moderator-2 Name', 'Login ID', 'Password (4-digit)', 'Mobile', 'Email ID',
        'Role', 'Day', 'Track Session', 'Room',
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row = 2
    for profile in profiles:
        email = matcher.find_email(profile.display_name) or ''
        phone = profile.phone or matcher.find_mobile(profile.display_name) or ''
        duties = get_faculty_duties(profile)
        if not duties:
            ws.cell(row, 1, profile.display_name)
            ws.cell(row, 2, profile.user.username)
            ws.cell(row, 3, profile.plain_password)
            ws.cell(row, 4, phone)
            ws.cell(row, 5, email)
            row += 1
            continue

        for item in duties:
            duty = item['duty']
            ws.cell(row, 1, profile.display_name)
            ws.cell(row, 2, profile.user.username)
            ws.cell(row, 3, profile.plain_password)
            ws.cell(row, 4, phone)
            ws.cell(row, 5, email)
            ws.cell(row, 6, ', '.join(item['roles']))
            ws.cell(row, 7, f'Day {duty.day}')
            ws.cell(row, 8, duty.track_session)
            ws.cell(row, 9, duty.room)
            for col in range(1, 10):
                ws.cell(row, col).border = border
                ws.cell(row, col).alignment = Alignment(vertical='center', wrap_text=True)
            row += 1

    widths = [28, 18, 18, 16, 28, 30, 10, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
