import re

from openpyxl import load_workbook


def _extract_day_info(sheet_name):
    name_upper = sheet_name.upper()
    if 'DAY 1' in name_upper or '12 JUNE' in name_upper:
        return 1, sheet_name
    if 'DAY 2' in name_upper or '13 JUNE' in name_upper:
        return 2, sheet_name
    return None, sheet_name


def _is_poster_serial(sr_str):
    return bool(re.match(r'^\d+_P$', str(sr_str or '').strip(), re.I))


def _is_poster_id(value):
    if value is None:
        return False
    text = str(value).strip()
    if text.upper() == 'POSTER':
        return False
    return bool(re.match(r'^[Pp][Oo][Ss][Tt][Ee][Rr]_?\d+', text))


def _is_poster_title(value):
    if not value:
        return False
    return 'POSTER PRESENTATION' in str(value).strip().upper()


def _is_poster_block(header_part, track_name_col):
    text = f'{header_part} {track_name_col}'.upper()
    return 'POSTER' in text


def _is_poster_paper_row(sr_str, track_session):
    if _is_poster_serial(sr_str):
        return True
    if track_session and str(track_session).strip().upper() == 'POSTER':
        return True
    return _is_poster_id(track_session)


def _room_from_sr(sr_str):
    if not sr_str:
        return ''
    match = re.match(r'^(\S+)', str(sr_str).strip())
    return match.group(1) if match else ''


def _resolve_poster_room(room, sr_str=''):
    """Pick display room for poster block (e.g. 503), not labels like S(I)_day1."""
    room_str = str(room or '').strip()
    if room_str.isdigit():
        return room_str
    match = re.search(r'\b(\d{2,4})\b', f'{room_str} {sr_str}')
    if match:
        return match.group(1)
    if room_str and not re.search(r'day\s*\d+', room_str, re.I):
        return room_str
    return ''


def _is_poster_section_header(sr_str, room, track_session, paper_title):
    """Brown header rows: POSTER PRESENTATION in merged cells, SR like S(I)_day1."""
    if _is_poster_serial(sr_str):
        return False
    if _is_poster_title(track_session) or _is_poster_title(paper_title) or _is_poster_title(room):
        return True
    row_text = f'{sr_str} {room} {track_session} {paper_title}'.upper()
    if 'POSTER PRESENTATION' in row_text:
        return True
    if re.search(r'day\s*\d+', str(sr_str or ''), re.I) and 'POSTER' in row_text:
        return True
    return False


def normalize_poster_room(value):
    """Normalize poster venue labels to match Paper.room (e.g. AUDI_ARCHITECTURE)."""
    text = str(value or '').strip()
    if not text:
        return ''
    if text.isdigit():
        return text
    text = re.sub(r'\s*poster\s*$', '', text, flags=re.I).strip()
    text = re.sub(r'[^a-zA-Z0-9]+', '_', text).strip('_')
    return text


def _activate_poster_session(room, sr_str='', existing_room='', label_room=''):
    session_label = 'POSTER PRESENTATION'
    for candidate in (room, label_room, existing_room):
        if not candidate:
            continue
        poster_room = _resolve_poster_room(candidate, sr_str) or normalize_poster_room(candidate)
        if poster_room:
            if not str(poster_room).isdigit():
                poster_room = normalize_poster_room(poster_room) or poster_room
            return (
                poster_room,
                session_label,
                poster_display_name(poster_room, session_label),
            )
    return ('', session_label, session_label)


def _poster_session_key(header_part, track_name_col):
    """Session key stored on Paper — must match TrackDuty.track_session."""
    if header_part and re.search(r'poster', str(header_part), re.I):
        if 'presentation' in str(header_part).lower():
            return str(header_part).strip()
        return str(header_part).strip()
    if track_name_col and 'POSTER' in str(track_name_col).upper():
        return str(track_name_col).strip()
    return 'POSTER PRESENTATION'


def poster_display_name(room, session_label):
    """Display label e.g. '503 | POSTER PRESENTATION'."""
    if room and session_label:
        return f'{room} | {session_label}'
    return session_label or room or ''


def _parse_track_session(track_session):
    if not track_session:
        return None, None, None
    text = str(track_session).strip()

    match = re.match(r'[Tt][Rr][Aa][Cc][Kk]\s*(\d+)_(\d+)', text)
    if match:
        track_num = int(match.group(1))
        subsession = int(match.group(2))
        display = f'TRACK {track_num:02d}_{subsession:02d}'
        return track_num, subsession, display

    poster_match = re.match(r'[Pp][Oo][Ss][Tt][Ee][Rr]_?(\d+)', text)
    if poster_match:
        poster_num = int(poster_match.group(1))
        display = f'POSTER_{poster_num:02d}'
        return poster_num, 0, display

    if text.upper() == 'POSTER':
        return 0, 0, 'POSTER'

    if 'POSTER' in text.upper():
        return 0, 0, text.upper().replace(' ', '_')

    return None, None, None


def _parse_serial(sr_value):
    if sr_value is None:
        return None, 0
    sr_str = str(sr_value).strip()
    if not sr_str or sr_str.upper() in ('SR', 'TIME'):
        return None, 0
    if _is_poster_serial(sr_str):
        order = int(re.match(r'^(\d+)', sr_str).group(1))
        return sr_str, order
    match = re.match(r'^(\d+)', sr_str)
    if not match:
        return None, 0
    return sr_str, int(match.group(1))


def generate_paper_id(track_session, day_num, serial_number, serial_order=None):
    text = str(track_session or '').strip()
    track_num, subsession, display = _parse_track_session(track_session)
    if track_num is None:
        return '', ''

    _, order = _parse_serial(serial_number)
    if serial_order is None:
        serial_order = order

    if text.upper() == 'POSTER' or (
        re.match(r'^[Pp][Oo][Ss][Tt][Ee][Rr]$', text) and serial_order
    ):
        paper_id = f'POSTER_{serial_order:02d}_D{day_num}_{serial_order:02d}'
        return paper_id, f'POSTER_{serial_order:02d}'

    if re.match(r'[Pp][Oo][Ss][Tt][Ee][Rr]', text):
        poster_match = re.match(r'[Pp][Oo][Ss][Tt][Ee][Rr]_?(\d+)', text)
        poster_num = int(poster_match.group(1)) if poster_match else serial_order or track_num
        paper_id = f'POSTER_{poster_num:02d}_D{day_num}_{serial_order:02d}'
        return paper_id, f'POSTER_{poster_num:02d}'

    paper_id = f'TRACK {track_num:02d}_{subsession:02d}_D{day_num}_{serial_order:02d}'
    return paper_id, display


def _apply_pipe_header(parts, track_session, current_room):
    current_room = parts[0].strip()
    header_track = parts[1].strip() if len(parts) > 1 else ''
    current_track_session = ''
    current_track_name = ''

    if re.match(r'[Tt]rack\s*\d+_\d+', header_track):
        current_track_session = header_track
        if track_session and 'Track' in str(track_session):
            current_track_name = str(track_session).strip()
    elif _is_poster_block(header_track, track_session):
        session_label = _poster_session_key(header_track, track_session)
        current_track_session = session_label
        current_track_name = poster_display_name(current_room, session_label)
    elif header_track:
        current_track_session = header_track
        if track_session:
            current_track_name = str(track_session).strip()
    elif track_session:
        current_track_session = str(track_session).strip()
        current_track_name = str(track_session).strip()

    return current_room, current_track_session, current_track_name


def _apply_poster_title_header(sr_str, room, paper_title, track_session='', existing_room=''):
    """Header rows: POSTER PRESENTATION in any column, SR like 503 day1 or S(I)_day1."""
    current_room, current_track_session, current_track_name = _activate_poster_session(
        room, sr_str, existing_room, label_room=track_session or paper_title
    )
    return current_room, current_track_session, current_track_name


def parse_schedule_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    papers = []
    tracks = set()

    for sheet_name in workbook.sheetnames:
        day_num, day_label = _extract_day_info(sheet_name)
        if day_num is None:
            continue

        worksheet = workbook[sheet_name]
        current_track_session = ''
        current_track_name = ''
        current_room = ''
        current_session_chair = ''

        for row in range(3, worksheet.max_row + 1):
            sr = worksheet.cell(row, 1).value
            room = worksheet.cell(row, 2).value
            track_session = worksheet.cell(row, 3).value
            paper_title = worksheet.cell(row, 4).value
            author_name = worksheet.cell(row, 5).value
            university = worksheet.cell(row, 6).value or ''
            mode = worksheet.cell(row, 7).value or ''
            session_chair = worksheet.cell(row, 8).value or ''
            time_slot = worksheet.cell(row, 13).value or ''

            sr_str = str(sr or '').strip()

            # Section header rows (with | or POSTER PRESENTATION title row)
            if sr_str and '|' in sr_str:
                current_room, current_track_session, current_track_name = _apply_pipe_header(
                    sr_str.split('|', 1), track_session, current_room
                )
                if session_chair:
                    current_session_chair = str(session_chair).strip()
                if current_track_session:
                    tracks.add((current_track_session, current_track_name, day_num))
                continue

            if _is_poster_section_header(sr_str, room, track_session, paper_title):
                current_room, current_track_session, current_track_name = _apply_poster_title_header(
                    sr_str, room, paper_title, track_session, current_room
                )
                if session_chair:
                    current_session_chair = str(session_chair).strip()
                tracks.add((current_track_session, current_track_name, day_num))
                continue

            serial_number, serial_order = _parse_serial(sr)
            if serial_number is None or not paper_title or not author_name:
                continue

            if _is_poster_title(paper_title):
                continue

            is_poster_paper = _is_poster_paper_row(sr_str, track_session)

            if track_session and re.match(r'[Tt]rack\s*\d+_\d+', str(track_session)):
                current_track_session = str(track_session).strip()
            elif is_poster_paper:
                poster_room = _resolve_poster_room(room) or str(room or '').strip()
                if poster_room:
                    current_room = poster_room
                current_track_session = 'POSTER PRESENTATION'
                current_track_name = poster_display_name(
                    current_room, 'POSTER PRESENTATION'
                )

            if room and not is_poster_paper:
                current_room = str(room).strip()
            elif room and is_poster_paper:
                poster_room = _resolve_poster_room(room)
                if poster_room:
                    current_room = poster_room
            if session_chair:
                current_session_chair = str(session_chair).strip()

            if not current_track_session:
                continue

            if is_poster_paper and (
                not track_session
                or str(track_session).strip().upper() == 'POSTER'
            ):
                paper_id, track_display = generate_paper_id(
                    'POSTER', day_num, serial_number, serial_order=serial_order
                )
            elif is_poster_paper and _is_poster_id(track_session):
                paper_id, track_display = generate_paper_id(
                    str(track_session).strip(), day_num, serial_number
                )
            else:
                paper_id, track_display = generate_paper_id(
                    current_track_session, day_num, serial_number
                )

            if not paper_id:
                continue

            tracks.add((current_track_session, current_track_name, day_num))

            papers.append({
                'day': day_num,
                'day_label': day_label,
                'serial_number': serial_number,
                'serial_order': serial_order,
                'room': current_room,
                'track_session': current_track_session,
                'track_name': current_track_name,
                'paper_title': str(paper_title).strip(),
                'author_name': str(author_name).strip(),
                'university': str(university).strip(),
                'mode': str(mode).strip(),
                'session_chair': current_session_chair,
                'time_slot': str(time_slot).strip() if time_slot else '',
                'paper_id': paper_id,
                'track_session_display': track_display,
            })

    workbook.close()
    return papers, tracks
