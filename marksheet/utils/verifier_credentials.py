import random
import re

from django.contrib.auth import get_user_model

from marksheet.models import TrackDuty, VerifierProfile
from marksheet.utils.faculty_credentials import get_faculty_duties, names_match
from marksheet.utils.faculty_matcher import FacultyMatcher, name_tokens, normalize_name

User = get_user_model()


def _generate_password():
    return f'{random.randint(1000, 9999):04d}'


def _generate_username(display_name, used_usernames):
    tokens = name_tokens(display_name)
    if len(tokens) >= 2:
        base = f'{tokens[0].lower()}.{tokens[-1].lower()}'
    else:
        base = re.sub(r'[^a-z0-9]', '', display_name.lower())[:20] or 'verifier'
    base = re.sub(r'[^a-z0-9.]', '', base)[:24] or 'verifier'
    username = base
    counter = 1
    while username in used_usernames or User.objects.filter(username=username).exists():
        suffix = str(counter)
        username = f'{base[:24 - len(suffix)]}{suffix}'
        counter += 1
    used_usernames.add(username)
    return username


def verifier_matches_name(profile, verifier_name):
    if not verifier_name:
        return False
    if names_match(profile.display_name, verifier_name):
        return True
    norm_profile = profile.normalized_name
    norm_field = normalize_name(verifier_name)
    if norm_profile == norm_field:
        return True
    if len(norm_field) <= 4 and norm_field in norm_profile.replace(' ', ''):
        return True
    profile_tokens = norm_profile.split()
    field_tokens = norm_field.split()
    if profile_tokens and field_tokens:
        if profile_tokens[0].startswith(field_tokens[0][:1]) and profile_tokens[-1].startswith(field_tokens[-1][:1]):
            if len(field_tokens) == 1 and len(norm_field) <= 4:
                initials = ''.join(t[0] for t in profile_tokens)
                return initials == norm_field
    return False


def get_verifier_duties(profile, schedule=None, day=None):
    if schedule is None:
        schedule = profile.schedule
    if not schedule:
        return []

    duties = TrackDuty.objects.filter(schedule=schedule).exclude(verifier='')
    if day:
        duties = duties.filter(day=day)

    matched = []
    for duty in duties:
        if verifier_matches_name(profile, duty.verifier):
            matched.append({'duty': duty, 'roles': ['Verifier']})
    return matched


def get_verifier_track_keys(profile, schedule=None, day=None):
    keys = set()
    for item in get_verifier_duties(profile, schedule=schedule, day=day):
        duty = item['duty']
        keys.add((duty.day, duty.track_session))
    return keys


def get_verifier_profiles(schedule):
    if not schedule:
        return VerifierProfile.objects.none()
    return VerifierProfile.objects.filter(schedule=schedule).select_related('user').order_by('display_name')


def sync_verifier_users(schedule):
    faculty_matcher = FacultyMatcher()
    unique_names = {}

    for duty in TrackDuty.objects.filter(schedule=schedule).exclude(verifier=''):
        name = duty.verifier.strip()
        norm = normalize_name(name) or name.upper()
        if norm not in unique_names:
            unique_names[norm] = name

    old_profiles = VerifierProfile.objects.filter(schedule=schedule).select_related('user')
    old_user_ids = [p.user_id for p in old_profiles]
    VerifierProfile.objects.filter(schedule=schedule).delete()
    User.objects.filter(id__in=old_user_ids, is_superuser=False).delete()

    used_usernames = set(User.objects.values_list('username', flat=True))
    created = []

    for norm, display_name in sorted(unique_names.items(), key=lambda x: x[1]):
        username = _generate_username(display_name, used_usernames)
        password = _generate_password()
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=display_name[:30],
        )
        phone = faculty_matcher.find_mobile(display_name) or ''
        profile = VerifierProfile.objects.create(
            user=user,
            display_name=display_name,
            normalized_name=norm,
            plain_password=password,
            phone=phone or '',
            schedule=schedule,
        )
        created.append(profile)

    return created


def generate_verifier_credentials_workbook(profiles):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Verifier Credentials'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Verifier Name', 'Login ID', 'Password (4-digit)', 'Mobile',
        'Role', 'Day', 'Track Session', 'Room',
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    for profile in profiles:
        duties = get_verifier_duties(profile)
        for item in duties or [{'duty': None, 'roles': ['Verifier']}]:
            duty = item.get('duty')
            ws.cell(row, 1, profile.display_name)
            ws.cell(row, 2, profile.user.username)
            ws.cell(row, 3, profile.plain_password)
            ws.cell(row, 4, profile.phone)
            ws.cell(row, 5, 'Verifier')
            if duty:
                ws.cell(row, 6, f'Day {duty.day}')
                ws.cell(row, 7, duty.track_session)
                ws.cell(row, 8, duty.room)
            row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
