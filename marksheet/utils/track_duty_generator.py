import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_parser import (
    _activate_poster_session,
    _extract_day_info,
    _is_poster_block,
    _is_poster_section_header,
    poster_display_name,
)
from .faculty_matcher import FacultyMatcher
from .track_colors import get_track_row_fill


HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
TITLE_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBTITLE_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
WHITE_BOLD = Font(bold=True, color='FFFFFF', size=12)
TITLE_FONT = Font(bold=True, color='FFFFFF', size=14)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(size=12, color='000000')
DATA_FONT_BOLD = Font(size=12, bold=True, color='000000')
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _format_room_cell(room, track_session):
    if track_session:
        return f'{room} - {track_session}'
    return room


def _format_track_cell(track_name, track_session):
    if track_name and track_name.upper() != str(track_session).upper():
        return track_name
    return track_session or track_name


def parse_track_duty(file_path):
    """Parse track duty rows (room, staff, moderators) per day from schedule Excel."""
    workbook = load_workbook(file_path, data_only=True)
    duty_by_day = {}

    for sheet_name in workbook.sheetnames:
        day_num, day_label = _extract_day_info(sheet_name)
        if day_num is None:
            continue

        worksheet = workbook[sheet_name]
        tracks = []

        for row in range(3, worksheet.max_row + 1):
            sr = worksheet.cell(row, 1).value
            sr_str = str(sr or '').strip()
            if not sr_str:
                continue

            track_name = worksheet.cell(row, 3).value
            track_name = str(track_name).strip() if track_name else ''
            paper_title = worksheet.cell(row, 4).value
            room_col = worksheet.cell(row, 2).value

            if _is_poster_section_header(sr_str, room_col, track_name, paper_title):
                room, track_session, track_name = _activate_poster_session(room_col, sr_str)
                track_session = 'POSTER PRESENTATION'
                session_chair = worksheet.cell(row, 8).value or ''
                track_coordinator = worksheet.cell(row, 9).value or ''
                moderator_1 = worksheet.cell(row, 10).value or ''
                moderator_2 = worksheet.cell(row, 11).value or ''
                moderator_3 = worksheet.cell(row, 12).value or ''
                tracks.append({
                    'room': room,
                    'track_session': track_session,
                    'track_name': track_name,
                    'room_track': track_name,
                    'session_chair': str(session_chair).strip(),
                    'track_coordinator': str(track_coordinator).strip(),
                    'moderator_1': str(moderator_1).strip(),
                    'moderator_2': str(moderator_2).strip(),
                    'moderator_3': str(moderator_3).strip(),
                })
                continue

            if '|' not in sr_str:
                continue

            parts = [p.strip() for p in sr_str.split('|')]
            room = parts[0]
            track_session = parts[1] if len(parts) > 1 else ''

            if not track_session and track_name:
                track_session = track_name

            if _is_poster_block(track_session, track_name):
                session_label = track_session if track_session and re.search(r'poster', track_session, re.I) else track_name
                track_name = poster_display_name(room, session_label or track_session)

            session_chair = worksheet.cell(row, 8).value or ''
            track_coordinator = worksheet.cell(row, 9).value or ''
            moderator_1 = worksheet.cell(row, 10).value or ''
            moderator_2 = worksheet.cell(row, 11).value or ''
            moderator_3 = worksheet.cell(row, 12).value or ''

            room_track = f'{room} | {track_session}' if track_session else room

            tracks.append({
                'room': room,
                'track_session': track_session,
                'track_name': track_name,
                'room_track': room_track,
                'session_chair': str(session_chair).strip(),
                'track_coordinator': str(track_coordinator).strip(),
                'moderator_1': str(moderator_1).strip(),
                'moderator_2': str(moderator_2).strip(),
                'moderator_3': str(moderator_3).strip(),
            })

        duty_by_day[day_num] = {
            'day_label': day_label,
            'tracks': tracks,
        }

    workbook.close()
    return duty_by_day


def _day_title_parts(day_label, day_num):
    match = re.search(r'(\d{1,2}\s+\w+\s+\d{4})', day_label or '', re.I)
    date_part = match.group(1).upper() if match else ''
    return day_num, date_part


def _fill_track_duty_sheet(ws, day_num, day_label, tracks, faculty_matcher=None):
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = 'ICRAET 2026  ◆  CONFERENCE PARALLEL TRACK SCHEDULE'
    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = CENTER
    title.border = BORDER

    _, date_part = _day_title_parts(day_label, day_num)
    ws.merge_cells('A2:G2')
    subtitle = ws['A2']
    subtitle.value = f'◆  DAY {day_num}  ◆  {date_part}'
    subtitle.font = WHITE_BOLD
    subtitle.fill = SUBTITLE_FILL
    subtitle.alignment = CENTER
    subtitle.border = BORDER

    headers = [
        'ROOM',
        'TRACK',
        'SESSION CHAIR',
        'TRACK COORDINATOR',
        'MODERATOR-1\n(ANCHOR)',
        'MODERATOR-2\n(ENTRY)',
        'MODERATOR-3\n(ZOOM)',
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for idx, track in enumerate(tracks, start=4):
        room_display = _format_room_cell(track['room'], track['track_session'])
        track_display = _format_track_cell(track['track_name'], track['track_session'])

        row_fill = get_track_row_fill(track['track_session'], track['track_name'])

        coordinator_fields = [
            track['track_coordinator'],
            track['moderator_1'],
            track['moderator_2'],
            track['moderator_3'],
        ]
        if faculty_matcher:
            coordinator_fields = [
                faculty_matcher.format_with_mobile(name)
                for name in coordinator_fields
            ]

        row_data = [
            room_display,
            track_display,
            track['session_chair'],
            *coordinator_fields,
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(idx, col, value)
            cell.fill = row_fill
            cell.font = DATA_FONT_BOLD if col <= 2 else DATA_FONT
            cell.alignment = CENTER if col == 1 else LEFT
            cell.border = BORDER

    col_widths = [28, 42, 28, 24, 22, 22, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 36
    for row in range(4, 4 + len(tracks)):
        ws.row_dimensions[row].height = 52 if faculty_matcher else 42

    ws.print_area = f'A1:G{3 + len(tracks)}'
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def generate_track_duty_workbook(duty_by_day, days=None, faculty_matcher=None):
    if days is None:
        days = sorted(duty_by_day.keys())

    if faculty_matcher is None:
        faculty_matcher = FacultyMatcher()

    workbook = Workbook()
    workbook.remove(workbook.active)

    for day_num in days:
        if day_num not in duty_by_day:
            continue
        day_info = duty_by_day[day_num]
        sheet_name = f'Day {day_num}'
        ws = workbook.create_sheet(title=sheet_name[:31])
        _fill_track_duty_sheet(
            ws,
            day_num,
            day_info['day_label'],
            day_info['tracks'],
            faculty_matcher=faculty_matcher,
        )

    if not workbook.sheetnames:
        ws = workbook.create_sheet('Empty')
        ws['A1'] = 'No track duty data found'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
