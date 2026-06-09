import re
from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties


LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_CENTER = Alignment(horizontal='right', vertical='center', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)


FIELD_MAP = {
    'paper id': 'paper_id',
    'paper title': 'paper_title',
    'author(s)': 'author_name',
    'session chair  name:': 'session_chair',
    'session chair name:': 'session_chair',
    'affiliation': 'university',
    'track / session': 'track_session_display',
}


def get_default_template_path():
    return Path(settings.BASE_DIR) / 'EVALUATION FORM (1).xlsx'


def resolve_template_path(template_file=None):
    if template_file:
        path = Path(template_file)
        if path.exists():
            return path
    default = get_default_template_path()
    if default.exists():
        return default
    return None


def _safe_sheet_name(name, used_names):
    invalid = re.compile(r'[\\/*?:\[\]]')
    cleaned = invalid.sub('', str(name)).strip()[:31]
    if not cleaned:
        cleaned = 'Sheet'
    base = cleaned
    counter = 1
    while cleaned in used_names:
        suffix = f'_{counter}'
        cleaned = base[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(cleaned)
    return cleaned


def _get_last_content_row(worksheet, max_col=4):
    for row in range(worksheet.max_row, 0, -1):
        for col in range(1, max_col + 1):
            value = worksheet.cell(row, col).value
            if value is not None and str(value).strip():
                return row
    return worksheet.max_row


def _apply_row_heights(worksheet, last_row=30):
    """Set row heights so the form fills a full A4 portrait page."""
    row_heights = {
        1: 30,    # Title
        2: 32,    # Paper ID
        3: 42,    # Paper Title (taller for long titles)
        4: 32,    # Author(s)
        5: 32,    # Session chair
        6: 32,    # Affiliation
        7: 32,    # Track / Session
        8: 24,    # Section A header
        9: 24,    # Table header
        10: 28, 11: 28, 12: 28, 13: 28, 14: 28,  # Criteria rows
        15: 26,   # Subtotal
        16: 24,   # Section B header
        17: 24,   # Table header
        18: 28, 19: 28, 20: 28, 21: 28, 22: 28,  # Criteria rows
        23: 26,   # Subtotal
        24: 24,   # Section C header
        25: 28,   # Final Score
        26: 88,   # Recommendation checkboxes
        27: 42,   # Comments (merged block)
        28: 42,
        29: 42,
        30: 36,   # Session Chair Sign
    }
    default_height = 28
    for row in range(1, last_row + 1):
        worksheet.row_dimensions[row].height = row_heights.get(row, default_height)


def _cell_text(worksheet, row, col=1):
    value = worksheet.cell(row, col).value
    return str(value).strip() if value is not None else ''


def _is_section_header_row(worksheet, row):
    text = _cell_text(worksheet, row, 1)
    return text.startswith('Section ')


def _is_table_header_row(worksheet, row):
    return _cell_text(worksheet, row, 1) == 'No.'


def _normalize_label(text):
    return ' '.join(str(text or '').strip().lower().split())


def _is_info_row(worksheet, row):
    key = _normalize_label(worksheet.cell(row, 1).value)
    if not key or 'comments' in key or 'sign' in key:
        return False
    info_labels = {
        'paper id', 'paper title', 'author(s)', 'affiliation', 'track / session',
        'session chair name:', 'session chair name',
    }
    return key in info_labels or key.startswith('session chair')


def _set_info_cell_value(worksheet, row, value):
    cell = worksheet.cell(row, 2)
    if not isinstance(cell, MergedCell):
        cell.value = value


LABEL_TO_FIELD = {
    'paper id': 'paper_id',
    'paper title': 'paper_title',
    'author(s)': 'author_name',
    'session chair name:': 'session_chair',
    'session chair name': 'session_chair',
    'track / session': 'track_session_display',
}


def fill_paper_data(worksheet, paper):
    """Fill paper info from schedule — Affiliation row stays blank."""
    for row in range(1, worksheet.max_row + 1):
        key = _normalize_label(worksheet.cell(row, 1).value)
        if not key:
            continue

        if key == 'affiliation':
            _set_info_cell_value(worksheet, row, None)
            continue

        field = LABEL_TO_FIELD.get(key)
        if not field and key.startswith('session chair'):
            field = 'session_chair'
        if not field:
            continue

        value = getattr(paper, field, '') or ''
        _set_info_cell_value(worksheet, row, value)


def _get_final_score_row(worksheet, last_row=30):
    for row in range(1, last_row + 1):
        for col in range(1, 5):
            if _cell_text(worksheet, row, col) == 'Final Score':
                return row
    return None


def _prepare_final_score_row(worksheet, last_row=30):
    row = _get_final_score_row(worksheet, last_row)
    if row and not _cell_text(worksheet, row, 3):
        cell = worksheet.cell(row, 3)
        if not isinstance(cell, MergedCell):
            cell.value = 'Out of (50)'
    return row


def _get_comments_block_rows(worksheet, last_row=30):
    for row in range(1, last_row + 1):
        for col in range(1, 5):
            text = _cell_text(worksheet, row, col).lower()
            if 'comments' in text and 'chair' in text:
                return {row, row + 1, row + 2}
    return {27, 28, 29}


def _apply_cell_alignments(worksheet, last_row=30):
    """Apply vertical center + horizontal left/center alignment across the form."""
    table_header_rows = set()
    for row in range(1, last_row + 1):
        if _is_table_header_row(worksheet, row):
            table_header_rows.add(row)

    evaluation_rows = set()
    comments_rows = _get_comments_block_rows(worksheet, last_row)
    final_score_row = _get_final_score_row(worksheet, last_row)
    for header_row in table_header_rows:
        for row in range(header_row + 1, last_row + 1):
            label = _cell_text(worksheet, row, 1)
            col_b = _cell_text(worksheet, row, 2)
            col_c = _cell_text(worksheet, row, 3)
            if label.lower() == 'subtotal' or col_c.lower() == 'subtotal':
                evaluation_rows.add(row)
                break
            if label.startswith('Section ') or label == 'No.':
                break
            if label.isdigit() or (label and label[0].isdigit()):
                evaluation_rows.add(row)
            elif col_b in ('Final Score',) or 'Recommendation' in col_b:
                break

    for row in range(1, last_row + 1):
        for col in range(1, 5):
            cell = worksheet.cell(row, col)

            # Title row
            if row == 1:
                cell.alignment = CENTER_CENTER
                continue

            # Section headers
            if _is_section_header_row(worksheet, row):
                cell.alignment = CENTER_CENTER
                continue

            # Table column headers
            if _is_table_header_row(worksheet, row):
                cell.alignment = CENTER_CENTER
                continue

            # Paper info rows — labels left, values left, vertically centered
            if _is_info_row(worksheet, row):
                cell.alignment = LEFT_CENTER
                continue

            # Evaluation criteria rows
            if row in evaluation_rows:
                if col == 2:
                    cell.alignment = LEFT_CENTER
                else:
                    cell.alignment = CENTER_CENTER
                continue

            # Final score row — Final Score right/center, Out of (50) center/center
            if row == final_score_row:
                text = _cell_text(worksheet, row, col)
                if text == 'Final Score':
                    cell.alignment = RIGHT_CENTER
                elif 'out of' in text.lower():
                    cell.alignment = CENTER_CENTER
                else:
                    cell.alignment = CENTER_CENTER
                continue

            # Recommendation row — left aligned, vertically centered
            if 'Recommendation' in _cell_text(worksheet, row, 2):
                cell.alignment = LEFT_CENTER
                continue

            # Session Chair Comments block — top-left aligned
            if row in comments_rows:
                cell.alignment = LEFT_TOP
                continue

            # Session Chair Sign — left aligned, vertically centered
            if 'sign' in _cell_text(worksheet, row, 1).lower():
                cell.alignment = LEFT_CENTER
                continue

            # Default: left horizontal, center vertical
            cell.alignment = LEFT_CENTER


def _apply_print_setup(worksheet):
    """Configure worksheet for full A4 portrait print preview on one page."""
    last_row = _get_last_content_row(worksheet)

    _apply_row_heights(worksheet, last_row)
    _prepare_final_score_row(worksheet, last_row)
    _apply_cell_alignments(worksheet, last_row)

    # Column widths tuned for A4 portrait (content area after punch margin)
    worksheet.column_dimensions['A'].width = 24.0
    worksheet.column_dimensions['B'].width = 50.0
    worksheet.column_dimensions['C'].width = 14.0
    worksheet.column_dimensions['D'].width = 14.0

    # Hide unused columns so print preview shows only the form
    for col_letter in 'EFGHIJKLMNOPQRSTUVWXYZ':
        worksheet.column_dimensions[col_letter].hidden = True

    worksheet.print_area = f'A1:D{last_row}'

    # Left margin widened for 3-hole punching (~1 inch / 2.5 cm)
    worksheet.page_margins.left = 1.0
    worksheet.page_margins.right = 0.45
    worksheet.page_margins.top = 0.55
    worksheet.page_margins.bottom = 0.45
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2

    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.verticalCentered = False
    worksheet.page_setup.scale = None

    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1

    worksheet.print_options.gridLines = False
    worksheet.print_options.horizontalCentered = False
    worksheet.sheet_view.view = 'pageBreakPreview'


def _copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def _copy_worksheet(source_ws, target_wb, sheet_title):
    target_ws = target_wb.create_sheet(title=sheet_title)

    for col_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dimension.width
        target_ws.column_dimensions[col_letter].hidden = dimension.hidden

    for row_num, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = dimension.height
        target_ws.row_dimensions[row_num].hidden = dimension.hidden

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, new_cell)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.print_options = copy(source_ws.print_options)

    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area

    return target_ws


def _get_template_sheet(workbook):
    if 'Evaluation Form' in workbook.sheetnames:
        return workbook['Evaluation Form']
    return workbook[workbook.sheetnames[0]]


def generate_marksheet_workbook(papers, sheet_name_fn=None, template_path=None):
    template_path = resolve_template_path(template_path)
    if not template_path:
        raise FileNotFoundError('No marksheet template found. Upload a template from the dashboard.')

    template_wb = load_workbook(template_path)
    template_ws = _get_template_sheet(template_wb)

    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    used_names = set()

    for index, paper in enumerate(papers, start=1):
        if sheet_name_fn:
            sheet_title = sheet_name_fn(paper, index)
        else:
            sheet_title = f'{paper.serial_order:02d}_{paper.paper_id}'
        sheet_title = _safe_sheet_name(sheet_title, used_names)
        new_ws = _copy_worksheet(template_ws, output_wb, sheet_title)
        fill_paper_data(new_ws, paper)
        _apply_print_setup(new_ws)

    template_wb.close()

    if not output_wb.sheetnames:
        empty = output_wb.create_sheet('Empty')
        empty['A1'] = 'No papers found'

    buffer = BytesIO()
    output_wb.save(buffer)
    buffer.seek(0)
    return buffer


def track_sheet_name(paper, index):
    return f'S{index:02d}_{paper.paper_id}'


def day_track_sheet_name(paper, index):
    return f'D{paper.day}_{paper.track_session}_{paper.serial_order:02d}'
