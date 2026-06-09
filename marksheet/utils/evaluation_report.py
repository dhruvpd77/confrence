from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .track_colors import get_track_row_fill

SECTION_A_FIELDS = [
    ('pres_clarity', 'Presentation Clarity'),
    ('originality', 'Originality & Creativity'),
    ('technical_knowledge', 'Technical Knowledge'),
    ('time_management', 'Time Management'),
    ('qa_handling', 'Q&A Handling'),
]

SECTION_B_FIELDS = [
    ('novelty', 'Novelty of Work'),
    ('methodology', 'Methodology'),
    ('result_validation', 'Result Validation'),
    ('impact', 'Impact / Application'),
    ('paper_quality', 'Paper Quality'),
]

# Column layout (1-based)
COL_STATUS = 9
COL_FINAL_SCORE = 10
COL_SEC_A_START = 16
COL_SEC_A_END = 20
COL_SEC_B_START = 21
COL_SEC_B_END = 25
COL_SEC_A_TOTAL = 26
COL_SEC_B_TOTAL = 27
COL_TOTAL_MARKS = 28
COL_COMMENTS = 29

DATA_FONT = Font(size=12)
DATA_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=12)
THIN = Side(style='thin', color='B0BEC5')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
FINAL_FILL = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')

STATUS_DONE_FONT = Font(size=12, bold=True, color='2E7D32')
STATUS_PENDING_FONT = Font(size=12, bold=True, color='E65100')
LOCKED_YES_FONT = Font(size=12, bold=True, color='C62828')
LOCKED_NO_FONT = Font(size=12, color='37474F')
TOTAL_FONT = Font(size=12, bold=True, color='1A237E')


def _fmt_dt(dt):
    if not dt:
        return ''
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime('%d %b %Y %H:%M')


def _evaluation_totals(ev):
    if not ev:
        return '', '', ''
    sec_a = ev.section_a_total
    sec_b = ev.section_b_total
    total = ev.final_score
    return (
        sec_a if sec_a is not None else '',
        sec_b if sec_b is not None else '',
        f'{total}/50' if total is not None else '',
    )


def build_evaluation_report_rows(papers, evaluations_by_paper_id, locks_map=None):
    locks_map = locks_map or {}
    rows = []
    for paper in papers:
        evaluation = evaluations_by_paper_id.get(paper.id)
        lock = locks_map.get((paper.day, paper.track_session))
        sec_a_total, sec_b_total, total_marks = _evaluation_totals(evaluation)
        rows.append({
            'paper': paper,
            'evaluation': evaluation,
            'status': 'Completed' if evaluation and evaluation.is_complete else 'Pending',
            'final_score': evaluation.final_score if evaluation else None,
            'section_a_total': sec_a_total,
            'section_b_total': sec_b_total,
            'total_marks': total_marks,
            'evaluator': (
                evaluation.evaluator.get_full_name() or evaluation.evaluator.username
                if evaluation and evaluation.evaluator else ''
            ),
            'recommendation': (
                evaluation.get_recommendation_display()
                if evaluation and evaluation.recommendation else ''
            ),
            'moderator_entered_at': _fmt_dt(evaluation.moderator_entered_at) if evaluation else '',
            'verifier_modified_at': _fmt_dt(evaluation.verifier_modified_at) if evaluation else '',
            'track_locked': 'Yes' if lock and lock.is_locked else 'No',
            'verifier_locked_at': _fmt_dt(lock.locked_at) if lock and lock.is_locked else '',
        })
    return rows


def _cell_font(col, value, status, track_locked):
    if col == COL_STATUS:
        return STATUS_DONE_FONT if value == 'Completed' else STATUS_PENDING_FONT
    if col == COL_TOTAL_MARKS or col == COL_FINAL_SCORE:
        if value not in ('', None):
            return TOTAL_FONT
    if col == 14:  # Track Locked
        return LOCKED_YES_FONT if value == 'Yes' else LOCKED_NO_FONT
    return DATA_FONT


def _cell_fill(col, track_fill, value):
    if col == COL_FINAL_SCORE and value not in ('', None):
        return FINAL_FILL
    if col in (COL_SEC_A_TOTAL, COL_SEC_B_TOTAL, COL_TOTAL_MARKS):
        return TOTAL_FILL
    return track_fill


def generate_evaluation_report_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluation Report'

    headers = [
        'Day', 'Track Session', 'Paper ID', 'Paper Title', 'Author', 'University', 'Mode',
        'Session Chair', 'Status', 'Final Score', 'Recommendation',
        'Moderator-2 Entry Time', 'Verifier Edit Time', 'Track Locked', 'Verifier Lock Time',
        'Pres. Clarity', 'Originality', 'Tech. Knowledge', 'Time Mgmt', 'Q&A',
        'Novelty', 'Methodology', 'Validation', 'Impact', 'Paper Quality',
        'Section A Total', 'Section B Total', 'Total Marks', 'Comments',
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = DATA_ALIGN

    for row_idx, item in enumerate(rows, start=2):
        paper = item['paper']
        ev = item['evaluation']

        track_fill = get_track_row_fill(paper.track_session, paper.track_name)

        values = [
            f'Day {paper.day}',
            paper.track_session,
            paper.paper_id,
            paper.paper_title,
            paper.author_name,
            paper.university,
            paper.mode,
            paper.session_chair,
            item['status'],
            item['final_score'] if item['final_score'] is not None else '',
            item['recommendation'],
            item['moderator_entered_at'],
            item['verifier_modified_at'],
            item['track_locked'],
            item['verifier_locked_at'],
        ]

        if ev:
            for field, _ in SECTION_A_FIELDS + SECTION_B_FIELDS:
                val = getattr(ev, field)
                values.append(val if val is not None else '')
            values.append(item['section_a_total'])
            values.append(item['section_b_total'])
            values.append(item['total_marks'])
            values.append(ev.comments or '')
        else:
            values.extend([''] * (len(SECTION_A_FIELDS) + len(SECTION_B_FIELDS) + 4))

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.border = BORDER
            cell.fill = _cell_fill(col, track_fill, value)
            cell.font = _cell_font(col, value, item['status'], item['track_locked'])
            cell.alignment = DATA_ALIGN

    col_widths = {
        1: 9, 2: 14, 3: 14, 4: 36, 5: 22, 6: 22, 7: 10, 8: 18,
        9: 12, 10: 12, 11: 18, 12: 20, 13: 20, 14: 12, 15: 20,
        16: 12, 17: 12, 18: 14, 19: 11, 20: 8,
        21: 10, 22: 13, 23: 12, 24: 10, 25: 13,
        26: 14, 27: 14, 28: 12, 29: 30,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}'
    ws.row_dimensions[1].height = 28

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def get_evaluations_map(evaluations):
    return {ev.paper_id: ev for ev in evaluations}
